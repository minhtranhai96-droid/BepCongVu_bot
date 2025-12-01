import os
import json
from datetime import datetime
import pytz
from flask import Flask, request
import telegram
from telegram import InlineKeyboardMarkup, InlineKeyboardButton
from fpdf import FPDF
from openpyxl import Workbook

TOKEN = os.getenv("BOT_TOKEN")
bot = telegram.Bot(token=TOKEN)

app = Flask(__name__)

DATA_FILE = "data.json"
VN_TIME = pytz.timezone("Asia/Ho_Chi_Minh")  # GMT+7


# ===================== FORMAT + PARSE TIỀN =====================

def format_money(amount):
    amount = int(amount)
    if amount >= 1_000_000:
        return f"{amount/1_000_000:.1f}M".rstrip("0").rstrip(".")
    elif amount >= 1_000:
        return f"{amount//1000}k"
    return str(amount)


def parse_money(text):
    text = text.lower().replace(" ", "").replace(",", ".")

    if text.endswith("k"):
        return int(float(text[:-1]) * 1000)

    if text.endswith("m"):
        return int(float(text[:-1]) * 1_000_000)

    if text.endswith("tr") or text.endswith("triệu"):
        return int(float(text.replace("tr", "").replace("triệu", "")) * 1_000_000)

    if text.endswith("ty") or text.endswith("tỷ"):
        return int(float(text[:-2]) * 1_000_000_000)

    return int(float(text))



# ===================== LƯU TRỮ DỮ LIỆU =====================

def load_data():
    if not os.path.exists(DATA_FILE):
        return {"quy": 0, "lich_su": []}
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)



# ===================== EXPORT FILE =====================

def generate_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "BaoCao"

    ws.append(["Thời gian", "Loại", "Số tiền", "Mô tả", "Người nhập"])

    for item in data["lich_su"]:
        ws.append([item["time"], item["type"], item["amount"], item["desc"], item["user"]])

    filename = f"Bao_cao_{datetime.now(VN_TIME).strftime('%Y%m')}.xlsx"
    wb.save(filename)
    return filename


def generate_pdf(data):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt="Báo cáo chi tiêu", ln=True, align='C')

    for item in data["lich_su"]:
        line = f"{item['time']} | {item['type']} | {format_money(item['amount'])} | {item['desc']} | {item['user']}"
        pdf.cell(0, 10, txt=line, ln=True)

    filename = f"Bao_cao_{datetime.now(VN_TIME).strftime('%Y%m')}.pdf"
    pdf.output(filename)
    return filename



# ===================== MENU BOT =====================

def send_menu(chat_id):
    buttons = [
        [InlineKeyboardButton("➕ Thêm quỹ", callback_data="add_quy")],
        [InlineKeyboardButton("➖ Chi tiêu", callback_data="chi_tieu")],
        [InlineKeyboardButton("📊 Báo cáo tổng hợp", callback_data="baocao")],
        [InlineKeyboardButton("📁 Xuất file", callback_data="export")],
    ]
    bot.send_message(chat_id, "Chọn chức năng:", reply_markup=InlineKeyboardMarkup(buttons))



# ===================== SERVER (WEBHOOK TELEGRAM) =====================

@app.route("/", methods=["GET"])
def home():
    return "BepCongVu Bot đang hoạt động!"

@app.route(f"/{TOKEN}", methods=["POST"])
def webhook():
    update = telegram.Update.de_json(request.get_json(), bot)

    # ===== XỬ LÝ NÚT BẤM =====
    if update.callback_query:
        chat_id = update.callback_query.message.chat_id
        user = update.callback_query.from_user.first_name
        action = update.callback_query.data

        if action == "add_quy":
            bot.send_message(chat_id, "Nhập số tiền muốn thêm (vd: 50k, 1m):")
            return "OK"

        elif action == "chi_tieu":
            bot.send_message(chat_id, "Nhập số tiền + mô tả (vd: 50k rau, 1m gas):")
            return "OK"

        elif action == "baocao":
            data = load_data()

            total_add = sum(i["amount"] for i in data["lich_su"] if i["type"] == "add")
            total_spend = sum(i["amount"] for i in data["lich_su"] if i["type"] == "spend")

            nap_list = [
                f"➕ {format_money(i['amount'])} — {i['desc']} ({i['user']}) • {i['time']}"
                for i in data["lich_su"] if i["type"] == "add"
            ]
            chi_list = [
                f"➖ {format_money(i['amount'])} — {i['desc']} ({i['user']}) • {i['time']}"
                for i in data["lich_su"] if i["type"] == "spend"
            ]

            nap_text = "\n".join(nap_list) if nap_list else "Không có"
            chi_text = "\n".join(chi_list) if chi_list else "Không có"

            msg = (
                "📊 **BÁO CÁO TỔNG HỢP**\n\n"
                f"💰 **Tổng nạp:** {format_money(total_add)}\n{nap_text}\n\n"
                f"🛒 **Tổng chi:** {format_money(total_spend)}\n{chi_text}\n\n"
                f"💵 **Quỹ còn lại:** {format_money(data['quy'])}"
            )

            bot.send_message(chat_id, msg, parse_mode="Markdown")
            return "OK"

        elif action == "export":
            data = load_data()
            excel = generate_excel(data)
            pdf = generate_pdf(data)

            bot.send_document(chat_id, open(excel, "rb"))
            bot.send_document(chat_id, open(pdf, "rb"))
            return "OK"


    # ===== XỬ LÝ NHẮN TIN =====
    if update.message:
        chat_id = update.message.chat_id
        txt = update.message.text
        user = update.message.from_user.first_name

        if txt.startswith("/start"):
            send_menu(chat_id)
            return "OK"

        # THÊM QUỸ
        if txt.replace(".", "").replace(",", "").replace("k", "").replace("m", "").replace("tr", "").replace("ty", "").replace("tỷ", "").isdigit():
            amount = parse_money(txt)

            data = load_data()
            data["quy"] += amount
            data["lich_su"].append({
                "time": datetime.now(VN_TIME).strftime("%d/%m/%Y %H:%M"),
                "type": "add",
                "amount": amount,
                "desc": "Nạp quỹ",
                "user": user
            })
            save_data(data)

            bot.send_message(chat_id, f"✔ Thêm {format_money(amount)} thành công!\n💰 Quỹ: {format_money(data['quy'])}")
            return "OK"

        # CHI TIÊU
        parts = txt.split(" ", 1)
        if len(parts) == 2:
            try:
                amount = parse_money(parts[0])
                desc = parts[1]
            except:
                bot.send_message(chat_id, "⚠ Sai định dạng! Ví dụ: `50k rau`, `1m gas`")
                return "OK"

            data = load_data()
            data["quy"] -= amount
            data["lich_su"].append({
                "time": datetime.now(VN_TIME).strftime("%d/%m/%Y %H:%M"),
                "type": "spend",
                "amount": amount,
                "desc": desc,
                "user": user
            })
            save_data(data)

            bot.send_message(chat_id, f"🧾 Chi {format_money(amount)} ({desc}) — bởi {user}\n💰 Còn: {format_money(data['quy'])}")
            return "OK"

    return "OK"
